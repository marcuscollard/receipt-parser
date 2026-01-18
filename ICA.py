#!/usr/bin/env python3
"""
Receipt Tracker – text/PDF, discounts, family allocation, suggestions, adjustments
-------------------------------------------------------------------------------
- Parse plain TEXT (paste receipts) or PDFs (optional) into detailed rows
- Fuzzy-match to your canonical items (aliases + family_pct)
- Apply per-item allocation (family vs personal) using `family_pct`
- Handle discounts: promo line + negative price applied to previous item
- Auto-suggest unknown items to items_suggested.csv
- Manual adjustments via a TSV file you can edit and re-run
- YAML config to avoid retyping flags

items.csv columns (CSV):
    canonical,aliases,family,family_pct
example:
    milk,"mjölk;mellanmjölk;helmjölk",0,0.60
    eggs,"ägg;egg",0,
    toilet paper,"toalettpapper;toapapper",1,
    picadeli sallad,"picadeli;sallad",0,0.25

adjustments.txt (TSV with header):
    # canonical_item    qty     unit_price_sek  total_price_sek family_pct  date    merchant    note
    milk    2                       44.00   0.70    2025-09-15  ICA Maxi    add milks
    toilet paper    1   19.90                   1.00            Willys      manual correction
    picadeli sallad 0.420           58.79   0.25                Coop        weight add

config.yaml (optional):
    month: dec  # or "all" to run all receipts_*.txt
    text_file: receipts.txt
    items: items.csv
    out: sammanfattning.csv
    report: kategoriserad.txt
    report_text: kategoriserad_rapport.txt
    report_xlsx: kategoriserad_rapport.xlsx
    adjust: adjustments.txt
    pdf_dir: null
    ocr: false
"""

from __future__ import annotations
import argparse
import csv
import glob
import os
import re
from dataclasses import dataclass
from datetime import datetime
from typing import List, Dict, Tuple, Optional

# Optional PDF stack (kept for completeness; text mode is preferred)
try:
    import pdfplumber  # type: ignore
except Exception:
    pdfplumber = None
try:
    import pytesseract  # type: ignore
    from pdf2image import convert_from_path  # type: ignore
    OCR_AVAILABLE = True
except Exception:
    OCR_AVAILABLE = False
    print("[WARN] pytesseract or pdf2image not installed; OCR will be disabled.")

try:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # type: ignore
    from openpyxl.utils.dataframe import dataframe_to_rows  # type: ignore
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False
    print("[WARN] openpyxl not installed; XLSX report output will be disabled.")

import pandas as pd  # type: ignore
from rapidfuzz import fuzz, process  # type: ignore
from dateutil import parser as dtparser  # type: ignore
import yaml  # type: ignore

# ---------- Regex helpers ----------
PRICE_RE = re.compile(r"(?<!\d)(\d{1,4}[.,]\d{2})(?:\s*kr)?\b", re.IGNORECASE)
QTY_PRICE_RE = re.compile(r"(?:(\d{1,3})\s*[x×]\s*)?(\d{1,4}[.,]\d{2})", re.IGNORECASE)
DATE_CAND_RE = re.compile(r"\b(\d{4}-\d{2}-\d{2}|\d{2}[./-]\d{2}[./-]\d{2,4})\b")
CURRENCY_CLEAN_RE = re.compile(r"\s*kr\b", re.IGNORECASE)
NEG_PRICE_RE = re.compile(r"^[\-\u2212]\s*(\d{1,4}[.,]\d{2})\s*kr\b", re.IGNORECASE)  # discounts

# ---------- Data model ----------
@dataclass
class ItemRow:
    canonical: str
    aliases: List[str]
    family_pct: float  # 0..1 default allocation to family

# ---------- IO helpers ----------
def read_textfile_text(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def read_pdf_text(path: str, use_ocr: bool = False) -> str:
    if not pdfplumber:
        return ""
    text = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
            text.append(t)
    joined = "\n".join(text)
    if use_ocr and (not joined.strip()) and OCR_AVAILABLE:
        images = convert_from_path(path)
        ocr_text = []
        for img in images:
            ocr_text.append(pytesseract.image_to_string(img, lang="swe+eng"))
        joined = "\n".join(ocr_text)
    return joined

# ---------- Items ----------
def load_items(items_csv: str) -> List[ItemRow]:
    items: List[ItemRow] = []
    with open(items_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            aliases = [a.strip() for a in (r.get("aliases") or "").split(";") if a.strip()]
            raw_pct = (r.get("family_pct") or "").strip()
            items.append(
                ItemRow(
                    canonical=r["canonical"].strip(),
                    aliases=aliases,
                    family_pct=raw_pct
                )
            )
    return items

def write_items_suggestions_from_textfile(text_file: str, items_csv: str, out_csv: Optional[str] = None) -> str:
    """
    Build a suggested items CSV by parsing a TEXT receipt file and collecting
    product names that do not match any existing canonical item or alias.

    Writes columns: canonical, aliases, family, family_pct
    - canonical: proposed canonical name (from receipt text)
    - aliases: empty (fill later)
    - family: 0 (fill later)
    - family_pct: empty like "" (fill later)
    """
    # Load existing items + aliases
    items = load_items(items_csv)
    terms, term_to_item = build_match_index(items)

    # Parse the text file using your existing parser
    text = read_textfile_text(text_file)
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    # Collect unique names from parsed line items (before matching)
    names = set()
    for raw_name, _price, _qty in parse_line_items(lines):
        base = re.sub(r"\s+", " ", raw_name).strip()
        if base:
            names.add(base)

    # Keep only names that don't match a canonical item (or any alias)
    unmatched = sorted(n for n in names if canonicalize(n, terms, term_to_item) is None)

    # Where to write suggestions
    if out_csv is None:
        out_csv = os.path.splitext(items_csv)[0] + "_suggested.csv"

    # Append (create with header if new)
    is_new_file = not os.path.exists(out_csv)
    with open(out_csv, "a", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        if is_new_file:
            w.writerow(["canonical", "aliases", "family", "family_pct"])
        for n in unmatched:
            w.writerow([n, "", 0, ""])

    print(f"[INFO] Suggested {len(unmatched)} new items -> {out_csv}")
    return out_csv


def build_match_index(items: List[ItemRow]) -> Tuple[List[str], Dict[str, ItemRow]]:
    terms: List[str] = []
    term_to_item: Dict[str, ItemRow] = {}
    for it in items:
        terms.append(it.canonical)
        term_to_item[it.canonical] = it
        for al in it.aliases:
            terms.append(al)
            term_to_item[al] = it
    return terms, term_to_item

def canonicalize(name: str, terms: List[str], term_to_item: Dict[str, ItemRow], threshold: int = 80) -> Optional[ItemRow]:
    low = name.lower()
    for t in terms:
        if t.lower() in low or low in t.lower():
            return term_to_item[t]
    best = process.extractOne(name, terms, scorer=fuzz.WRatio)
    if best and best[1] >= threshold:
        return term_to_item[best[0]]
    return None


def parse_line_items(lines: List[str]) -> List[Tuple[str, float, float]]:
    """
    Supports:
      Name -> 24,95 kr -> (optional) "3 st * 20,95 kr/st" or "0,705 kg * 139,95 kr/kg"
      Discounts: promo line followed by negative price (e.g., "−21,00 kr") applies to previous item.
    Returns: list of (raw_name, total_price_sek, qty) where qty may be float (kg).
    """
    def is_noise(s: str) -> bool:
        low = s.lower()
        noise = ["summa", "subtotal", "att betala", "totalt", "moms", "kvitt", "kvitto",
                 "tack", "återköp", "card", "visa", "mastercard", "swish", "faktura", "personal", "kampanj"]
        return any(k in low for k in noise)

    def extract_total(s: str) -> Optional[float]:
        m = PRICE_RE.search(s)
        if not m:
            return None
        try:
            return float(m.group(1).replace(",", "."))
        except Exception:
            return None

    items: List[Tuple[str, float, float]] = []
    i, n = 0, len(lines)
    while i < n:
        ln = (lines[i] or "").strip()
        if not ln or is_noise(ln):
            i += 1
            continue

        # Discount: description line followed by a negative amount line
        if i + 1 < n and NEG_PRICE_RE.search(lines[i + 1].strip()):
            if items:
                m = NEG_PRICE_RE.search(lines[i + 1].strip())
                disc = float(m.group(1).replace(",", ".")) if m else 0.0
                name_prev, total_prev, qty_prev = items[-1]
                items[-1] = (name_prev, round(total_prev - disc, 2), qty_prev)
            i += 2
            continue

        # Single-line: trailing price
        single_price_match = None
        for m in PRICE_RE.finditer(ln):
            single_price_match = m
        if single_price_match:
            price = float(single_price_match.group(1).replace(",", "."))
            name_part = ln[:single_price_match.start()].strip()
            name_part = CURRENCY_CLEAN_RE.sub("", name_part).strip() if name_part else "UNKNOWN ITEM"
            qty_val: float = 1.0
            qp = QTY_PRICE_RE.search(ln)
            if qp and qp.group(1):
                try:
                    qty_val = float(qp.group(1))
                except Exception:
                    qty_val = 1.0
            items.append((" ".join(name_part.split()), round(price, 2), qty_val))
            i += 1
            continue

        # Multiline: name line, next line price, optional third line detail
        name_ml = " ".join(ln.split())
        total = extract_total(lines[i + 1].strip()) if i + 1 < n else None
        if total is None:
            i += 1
            continue
        qty_val = 1.0
        consumed = 2
        if i + 2 < n:
            det = " ".join((lines[i + 2] or "").strip().lower().split())
            if (" st" in det and "kr/st" in det) or (" kg" in det and "kr/kg" in det) or (" st *" in det) or (" kg *" in det):
                nums = re.findall(r"\d+(?:[.,]\d+)?", det)
                if nums:
                    try:
                        qty_val = float(nums[0].replace(",", "."))
                    except Exception:
                        qty_val = 1.0
                consumed = 3
        items.append((name_ml, round(total, 2), qty_val))
        i += consumed

    return items


# ---------- Manual adjustments ----------
def load_adjustments(path: str) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    if not path or not os.path.exists(path):
        return rows
    with open(path, "r", encoding="utf-8") as f:
        rdr = csv.DictReader(
            (ln for ln in f if not ln.strip().startswith("#") and ln.strip()),
            delimiter="\t",
        )
        for r in rdr:
            canon = (r.get("canonical_item") or "").strip()
            if not canon:
                continue
            qty = float((r.get("qty") or "1").replace(",", "."))
            unit = r.get("unit_price_sek") or ""
            total = r.get("total_price_sek") or ""
            unit_val = float(unit.replace(",", ".")) if unit else None
            total_val = float(total.replace(",", ".")) if total else None
            if unit_val is None and total_val is None:
                continue
            if unit_val is None:
                unit_val = round(total_val / max(qty, 1e-9), 2)
            if total_val is None:
                total_val = round(unit_val * qty, 2)
            fam_pct = r.get("family_pct")
            fam_pct_val = float(fam_pct) if fam_pct and fam_pct.strip() != "" else None
            rows.append({
                "date": (r.get("date") or "").strip(),
                "receipt_file": os.path.basename(path),
                "canonical_item": canon,
                "matched_text": f"[manual] {(r.get('note') or '').strip()}",
                "qty": qty,
                "unit_price_sek": round(unit_val, 2),
                "total_price_sek": round(total_val, 2),
                "family_pct": fam_pct_val,
            })
    return rows

# ---------- Output path helpers ----------
def path_in_dir(path: str, out_dir: str) -> str:
    if not path or os.path.isabs(path) or os.path.dirname(path):
        return path
    return os.path.join(out_dir, path)

def add_month_suffix(path: str, month: str) -> str:
    if not path:
        return path
    base = os.path.basename(path)
    root, ext = os.path.splitext(base)
    suffix = f"_{month}"
    if root.endswith(suffix):
        return path
    new_base = f"{root}{suffix}{ext}"
    parent = os.path.dirname(path)
    return os.path.join(parent, new_base) if parent else new_base

def months_from_receipts() -> List[str]:
    months: List[str] = []
    for path in sorted(glob.glob("receipts_*.txt")):
        base = os.path.basename(path)
        if not (base.startswith("receipts_") and base.endswith(".txt")):
            continue
        month = base[len("receipts_"):-len(".txt")]
        if month:
            months.append(month)
    return months

def read_categorized_report(path: str) -> pd.DataFrame:
    with open(path, "r", encoding="utf-8") as f:
        header_line = f.readline()
        if not header_line:
            raise ValueError(f"Empty report: {path}")
        header = header_line.lstrip("#").strip()
        columns = [c.strip() for c in header.split("\t") if c.strip()]
        rows: List[List[str]] = []
        for line in f:
            if not line.strip() or line.lstrip().startswith("#"):
                continue
            rows.append(line.rstrip("\n").split("\t"))
    df = pd.DataFrame(rows, columns=columns)
    for col in columns:
        if col == "canonical":
            continue
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

def build_swedish_report_df(df: pd.DataFrame) -> pd.DataFrame:
    column_map = {
        "canonical": "vara",
        "qty": "antal",
        "avg_unit_price_sek": "snitt pris",
        "total_price_sek": "totalpris",
        "family_alloc_sek": "varav mormor",
        "effective_family_pct": "eff. procent",
    }
    available = [c for c in column_map if c in df.columns]
    swedish_df = df[available].rename(columns=column_map)
    if "varav mormor" in swedish_df.columns:
        swedish_df.sort_values("varav mormor", ascending=False, inplace=True, ignore_index=True)
    return swedish_df

def write_categorized_text_report(df: pd.DataFrame, out_path: str) -> str:
    if df.empty:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write("No categorized rows.\n")
        return out_path

    headers = list(df.columns)
    numeric_cols = {c for c in headers if pd.api.types.is_numeric_dtype(df[c])}
    format_map = {
        "antal": "{:.3f}",
        "snitt pris": "{:.2f}",
        "totalpris": "{:.2f}",
        "varav mormor": "{:.2f}",
        "eff. procent": "{:.2f}%",
    }

    formatted_rows: List[List[str]] = []
    for _, r in df.iterrows():
        row: List[str] = []
        for col in headers:
            val = r.get(col, "")
            if pd.isna(val):
                row.append("")
                continue
            if col == "eff. procent":
                row.append(format_map[col].format(float(val) * 100.0))
            elif col in format_map:
                row.append(format_map[col].format(float(val)))
            else:
                row.append(str(val))
        formatted_rows.append(row)

    totals = {col: "" for col in headers}
    if headers and any(col not in format_map for col in headers):
        totals[headers[0]] = "TOTAL"
    for col in ["antal", "totalpris", "varav mormor"]:
        if col in df.columns:
            totals[col] = format_map[col].format(float(df[col].sum()))
    formatted_rows.append([totals.get(col, "") for col in headers])

    widths: List[int] = []
    for i, col in enumerate(headers):
        max_len = len(str(col))
        for row in formatted_rows:
            max_len = max(max_len, len(str(row[i])))
        widths.append(max_len)

    def format_row(row: List[str]) -> str:
        cells: List[str] = []
        for i, cell in enumerate(row):
            if headers[i] in numeric_cols:
                cells.append(str(cell).rjust(widths[i]))
            else:
                cells.append(str(cell).ljust(widths[i]))
        return "  ".join(cells)

    sep = "-" * (sum(widths) + 2 * (len(widths) - 1))
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(format_row(headers) + "\n")
        f.write(sep + "\n")
        for row in formatted_rows:
            f.write(format_row(row) + "\n")

    return out_path

def write_categorized_xlsx_report(df: pd.DataFrame, out_path: str) -> Optional[str]:
    if not OPENPYXL_AVAILABLE:
        print("[WARN] openpyxl not installed; skipping XLSX report.")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="C9C9C9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    currency_fmt = "#,##0.00"
    qty_fmt = "0.000"
    pct_fmt = "0.00%"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        for c_idx, _ in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = border
            cell.alignment = align_left

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    col_to_idx = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    for col_name in ["antal", "snitt pris", "totalpris", "varav mormor", "eff. procent"]:
        if col_name in col_to_idx:
            c = col_to_idx[col_name]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = align_right
                if col_name == "antal":
                    cell.number_format = qty_fmt
                elif col_name == "eff. procent":
                    cell.number_format = pct_fmt
                else:
                    cell.number_format = currency_fmt

    totals_row = ws.max_row + 1
    first_header = ws.cell(row=1, column=1).value
    if first_header not in {"antal", "snitt pris", "totalpris", "varav mormor", "eff. procent"}:
        ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=totals_row, column=1).alignment = align_left

    for name in ["antal", "totalpris", "varav mormor"]:
        if name in col_to_idx:
            c = col_to_idx[name]
            col_letter = ws.cell(row=1, column=c).column_letter
            ws.cell(row=totals_row, column=c, value=f"=SUM({col_letter}2:{col_letter}{totals_row-1})")
            ws.cell(row=totals_row, column=c).font = Font(bold=True)
            ws.cell(row=totals_row, column=c).alignment = align_right
            ws.cell(row=totals_row, column=c).border = border
            ws.cell(row=totals_row, column=c).number_format = qty_fmt if name == "antal" else currency_fmt

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=totals_row, column=c)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="FFF7CC")
        if cell.value is None:
            cell.value = ""

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = max(12, len(str(header)) + 2)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    ws.sheet_view.showGridLines = False

    wb.save(out_path)
    return out_path

def run_with_args(args: argparse.Namespace) -> None:
    family_summary_path = "familj_sammanfattning.txt"
    if args.month:
        month = str(args.month).strip().lower()
        if month:
            out_dir = month
            os.makedirs(out_dir, exist_ok=True)
            if args.pdf_dir is None:
                if args.text_file is None or os.path.basename(args.text_file) == "receipts.txt":
                    args.text_file = f"receipts_{month}.txt"
            args.out = path_in_dir(args.out, out_dir)
            args.report = path_in_dir(args.report, out_dir)
            if args.report_text:
                args.report_text = path_in_dir(args.report_text, out_dir)
            if args.report_xlsx:
                args.report_xlsx = path_in_dir(args.report_xlsx, out_dir)
            family_summary_path = path_in_dir(family_summary_path, out_dir)
            args.out = add_month_suffix(args.out, month)
            args.report = add_month_suffix(args.report, month)
            if args.report_text:
                args.report_text = add_month_suffix(args.report_text, month)
            if args.report_xlsx:
                args.report_xlsx = add_month_suffix(args.report_xlsx, month)
            family_summary_path = add_month_suffix(family_summary_path, month)
        else:
            args.month = None

    # check if items file exists

    if args.overwrite:
        write_items_suggestions_from_textfile(args.text_file, args.items, out_csv="items.csv")

    items = load_items(args.items)
    items_idx = build_match_index(items)

    rows: List[Dict[str, object]] = []

    if args.text_file:
        text = read_textfile_text(args.text_file)
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        unmatched = set()
        terms, term_to_item = items_idx
        for raw_name, price, qty in parse_line_items(lines):
            it = canonicalize(raw_name, terms, term_to_item)
            if not it:
                base = re.sub(r"\s+", " ", raw_name).strip()
                if base and len(base) <= 80:
                    unmatched.add(base)
                continue
            rows.append({
                "canonical_item": it.canonical,
                "matched_text": raw_name,
                "qty": qty,
                "unit_price_sek": round(price / max(qty, 1), 2),
                "total_price_sek": round(price, 2),
                "family_pct": it.family_pct,
            })
        if unmatched:
            sugg_path = os.path.splitext(args.items)[0] + "_suggested.csv"
            new_file = not os.path.exists(sugg_path)
            with open(sugg_path, "a", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                if new_file:
                    w.writerow(["canonical", "aliases", "family", "family_pct"])
                for name in sorted(unmatched):
                    w.writerow([name, "", 0, ""])
            print(f"[INFO] Wrote suggestions for {len(unmatched)} new items to {sugg_path}")

    elif args.pdf_dir:
        if not pdfplumber:
            raise SystemExit("pdfplumber is not installed; use --text-file or install PDF deps.")
        for root, _, files in os.walk(args.pdf_dir):
            for fn in files:
                if not fn.lower().endswith(".pdf"):
                    continue
                path = os.path.join(root, fn)
                try:
                    rows.extend(process_pdf(path, items_idx, use_ocr=args.ocr))
                except Exception as e:
                    print(f"[WARN] Failed parsing {path}: {e}")
                    continue
    else:
        raise SystemExit("Provide either --text-file or --pdf-dir")

    # Merge manual adjustments
    if args.adjust:
        rows.extend(load_adjustments(args.adjust))

    if not rows:
        print("No matching items found. Check items.csv aliases or run once to get items_suggested.csv.")
        return

    # Detailed CSV
    df = pd.DataFrame(rows)

    # Fill per-row family_pct from items defaults if missing
    item_family_pct = {it.canonical: it.family_pct for it in items}
    if "family_pct" not in df.columns:
        df["family_pct"] = None
    df["family_pct"] = df.apply(
        lambda r: r["family_pct"] if pd.notnull(r.get("family_pct")) else item_family_pct.get(r["canonical_item"], 0.0),
        axis=1,
    ).astype(float)

    df["family_alloc_sek"] = (df["total_price_sek"] * df["family_pct"]).round(2)
    df["personal_alloc_sek"] = (df["total_price_sek"] * (1.0 - df["family_pct"])).round(2)

    df.sort_values(["family_alloc_sek", "canonical_item"], inplace=True, ascending=[False, True])
    df.to_csv(args.out, index=False, encoding="utf-8")

    # Given a pandas DataFrame `df` with at least:
    # ['canonical_item','qty','unit_price_sek','total_price_sek']
    # and EITHER per-row allocation columns:
    #   - 'family_alloc_sek' (preferred), or 'family_pct' (0..1), or a binary 'family' flag.
    # This script writes a TSV file with:
    # 1) family % of total, family spend, total spend
    # 2) family-only items sorted by total family cost.

    def write_family_summary_from_df(df: pd.DataFrame, out_path: str = "familj_sammanfattning.txt") -> str:
        # Ensure numeric types
        for col in ["qty", "unit_price_sek", "total_price_sek"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

        # Build family_alloc_sek if missing
        if "family_alloc_sek" not in df.columns:
            if "family_pct" in df.columns:
                df["family_pct"] = pd.to_numeric(df["family_pct"], errors="coerce").fillna(0.0).clip(0.0, 1.0)
                df["family_alloc_sek"] = (df["total_price_sek"] * df["family_pct"]).round(2)
            elif "family" in df.columns:
                df["family"] = pd.to_numeric(df["family"], errors="coerce").fillna(0.0)
                df["family_alloc_sek"] = (df["total_price_sek"] * (df["family"] > 0).astype(float)).round(2)
            else:
                # No family info at all → assume zero allocation
                df["family_alloc_sek"] = 0.0

        # Totals
        grand_total = float(df["total_price_sek"].sum()) if len(df) else 0.0
        family_total = float(df["family_alloc_sek"].sum()) if len(df) else 0.0
        family_pct_total = (family_total / grand_total) if grand_total > 0 else 0.0

        # Aggregate by canonical item for family-only ranking
        # If qty/unit_price are missing, fall back gracefully
        agg_map = {
            "qty": ("qty", "sum") if "qty" in df.columns else ("total_price_sek", "size"),
            "avg_unit_price_sek": ("unit_price_sek", "mean") if "unit_price_sek" in df.columns else ("total_price_sek", "mean"),
            "total_family_sek": ("family_alloc_sek", "sum"),
        }
        fam_items = (
            df.groupby("canonical_item")
            .agg(**{k: pd.NamedAgg(*v) for k, v in agg_map.items()})
            .reset_index()
        )
        fam_items = fam_items[fam_items["total_family_sek"] > 0].copy()
        fam_items.sort_values("total_family_sek", ascending=False, inplace=True)

        # Write TSV
        with open(out_path, "w", encoding="utf-8") as f:
            f.write("# summary\n")
            f.write(f"family_percentage_of_total\t{family_pct_total:.4f}\n")
            f.write(f"family_expenditures_sek\t{family_total:.2f}\n")
            f.write(f"total_expenditures_sek\t{grand_total:.2f}\n")
            f.write("\n# family_items_by_total_cost\n")
            f.write("canonical\tqty\tavg_unit_price_sek\ttotal_family_sek\n")
            for _, r in fam_items.iterrows():
                qty_val = float(r["qty"]) if pd.notnull(r["qty"]) else 0.0
                avg_val = float(r["avg_unit_price_sek"]) if pd.notnull(r["avg_unit_price_sek"]) else 0.0
                tot_val = float(r["total_family_sek"]) if pd.notnull(r["total_family_sek"]) else 0.0
                f.write(f"{r['canonical_item']}\t{qty_val:.3f}\t{avg_val:.2f}\t{tot_val:.2f}\n")

        return out_path

    out_file = write_family_summary_from_df(df, family_summary_path)

    # Categorized summary (TSV)
    agg = (
        df.groupby("canonical_item")
        .agg(
            qty=("qty", "sum"),
            avg_unit_price_sek=("unit_price_sek", "mean"),
            total_price_sek=("total_price_sek", "sum"),
            family_alloc_sek=("family_alloc_sek", "sum"),
            personal_alloc_sek=("personal_alloc_sek", "sum"),
        )
        .reset_index()
    )
    agg["effective_family_pct"] = (agg["family_alloc_sek"] / agg["total_price_sek"]).replace([pd.NA, float("inf")], 0.0).fillna(0.0).round(2)
    agg.sort_values("total_price_sek", ascending=False, inplace=True)

    with open(args.report, "w", encoding="utf-8") as f:
        f.write("# canonical\tqty\tavg_unit_price_sek\ttotal_price_sek\tfamily_alloc_sek\tpersonal_alloc_sek\teffective_family_pct\n")
        for _, r in agg.iterrows():
            f.write(
                f"{r['canonical_item']}\t{r['qty']}\t{r['avg_unit_price_sek']:.2f}\t"
                f"{r['total_price_sek']:.2f}\t{r['family_alloc_sek']:.2f}\t{r['personal_alloc_sek']:.2f}\t"
                f"{r['effective_family_pct']:.2f}\n"
            )

    if args.report_text or args.report_xlsx:
        categorized_df = read_categorized_report(args.report)
        report_df = build_swedish_report_df(categorized_df)
        if args.report_text:
            write_categorized_text_report(report_df, args.report_text)
        if args.report_xlsx:
            write_categorized_xlsx_report(report_df, args.report_xlsx)

    print("\nTop categories by total (SEK):")
    for _, r in agg.iterrows():
        print(f"- {r['canonical_item']}: x{r['qty']} avg {r['avg_unit_price_sek']:.2f} -> {r['total_price_sek']:.2f} SEK (family {r['effective_family_pct']:.0%})")

# ---------- Main ----------
def main():
    ap = argparse.ArgumentParser(description="Ingest receipts from TEXT or PDFs, apply discounts, family allocations, and summarize.")
    ap.add_argument("--pdf-dir", help="Folder containing receipt PDFs")
    ap.add_argument("--text-file", help="Path to a plain text file with pasted receipt text")
    ap.add_argument("--month", help="Shortcut: use receipts_<month>.txt and write outputs to ./<month>/ (or 'all')")
    ap.add_argument("--items", required=False, help="Path to items.csv (canonical, aliases, family, family_pct)")
    ap.add_argument("--out", default="summary.csv", help="Output CSV path (detailed rows)")
    ap.add_argument("--report", default="categorized.txt", help="Plain text summary (sorted by total desc)")
    ap.add_argument("--report-text", help="Formatted text report derived from categorized.txt")
    ap.add_argument("--report-xlsx", help="Spreadsheet report derived from categorized.txt (requires openpyxl)")
    ap.add_argument("--adjust", help="TSV file with manual rows to add/override after parsing")
    ap.add_argument("--ocr", action="store_true", help="Enable OCR fallback for image-only PDFs (if using --pdf-dir)")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing output files instead of appending suggestions")

    # YAML defaults
    if os.path.exists("config.yaml"):
        try:
            print("[INFO] Loading defaults from config.yaml")
            with open("config.yaml", "r", encoding="utf-8") as f:
                defaults = yaml.safe_load(f) or {}
            for k, v in defaults.items():
                if v is not None:
                    ap.set_defaults(**{k: v})
        except Exception as e:
            print(f"[WARN] Failed to read config.yaml: {e}")
    else:
        print("[INFO] No config.yaml found; using command-line args or defaults")

    args = ap.parse_args()

    month_value = str(args.month).strip().lower() if args.month is not None else ""
    if month_value == "all":
        months = months_from_receipts()
        if not months:
            raise SystemExit("No receipts_*.txt files found for month: all")
        for month in months:
            args_month = argparse.Namespace(**vars(args))
            args_month.month = month
            run_with_args(args_month)
        return

    run_with_args(args)

if __name__ == "__main__":
    main()
